from flask import Flask, render_template, request, redirect, url_for, jsonify, session
import os, hashlib, psycopg2, psycopg2.extras
from datetime import datetime, date, timedelta
from functools import wraps

# ── helpers ──────────────────────────────────────────────
def is_attendance_open(event_date_str):
    try:
        event_date = datetime.strptime(str(event_date_str), "%Y-%m-%d").date()
        today = date.today()
        return event_date <= today <= event_date + timedelta(days=2)
    except:
        return False

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

# Types that use satsangi table
SABHA_TYPES = {'sunday', 'wednesday', 'balsabha'}

def is_sabha_type(event_type):
    return event_type in SABHA_TYPES

def event_type_label(event_type):
    return {'sunday': 'Sunday Sabha', 'wednesday': 'Wednesday Sabha',
            'balsabha': 'Bal Sabha', 'hostel': 'Hostel Assembly'}.get(event_type, event_type.title())

def categorize_events(events_raw):
    today = date.today()
    active, upcoming, past = [], [], []
    for e in events_raw:
        try:
            edate = datetime.strptime(str(e['event_date']), '%Y-%m-%d').date()
        except:
            past.append(e); continue
        if edate <= today <= edate + timedelta(days=2):
            e['status'] = 'active'; active.append(e)
        elif edate > today:
            e['status'] = 'upcoming'; upcoming.append(e)
        else:
            e['status'] = 'past'; past.append(e)
    upcoming.sort(key=lambda x: str(x['event_date']))
    return active, upcoming, past

# ── Config ───────────────────────────────────────────────
ADMIN_ID   = os.environ.get('ADMIN_ID',   '1234')
ADMIN_PASS = hash_password(os.environ.get('ADMIN_PASS', '5005'))

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-only-insecure-key-change-in-prod')

DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

# ── DB ────────────────────────────────────────────────────
def get_db():
    return psycopg2.connect(DATABASE_URL)

def dict_row(cursor, row):
    cols = [desc[0] for desc in cursor.description]
    return dict(zip(cols, row))

def fetchall_dict(cursor):
    return [dict_row(cursor, r) for r in cursor.fetchall()]

def fetchone_dict(cursor):
    row = cursor.fetchone()
    return dict_row(cursor, row) if row else None

# ── INIT DB ───────────────────────────────────────────────
def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS buildings (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL, description TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS rooms (
        id SERIAL PRIMARY KEY,
        building_id INTEGER NOT NULL REFERENCES buildings(id) ON DELETE CASCADE,
        room_number TEXT NOT NULL, capacity INTEGER DEFAULT 4, floor INTEGER DEFAULT 1)''')
    c.execute('''CREATE TABLE IF NOT EXISTS students (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL, roll_number TEXT, phone TEXT,
        room_id INTEGER REFERENCES rooms(id) ON DELETE SET NULL, joining_date TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS satsangis (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL, mobile TEXT, address TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS events (
        id SERIAL PRIMARY KEY, title TEXT NOT NULL, event_date TEXT NOT NULL,
        event_type TEXT NOT NULL DEFAULT 'hostel',
        description TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    # Add event_type column if upgrading existing DB
    c.execute("""SELECT column_name FROM information_schema.columns
                 WHERE table_name='events' AND column_name='event_type'""")
    if not c.fetchone():
        c.execute("ALTER TABLE events ADD COLUMN event_type TEXT NOT NULL DEFAULT 'hostel'")
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id SERIAL PRIMARY KEY,
        event_id INTEGER NOT NULL REFERENCES events(id) ON DELETE CASCADE,
        person_id INTEGER NOT NULL,
        person_type TEXT NOT NULL DEFAULT 'student',
        status TEXT NOT NULL DEFAULT 'present',
        marked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(event_id, person_id, person_type))''')
    # Migrate old attendance table if needed
    c.execute("""SELECT column_name FROM information_schema.columns
                 WHERE table_name='attendance' AND column_name='person_type'""")
    if not c.fetchone():
        c.execute("ALTER TABLE attendance ADD COLUMN person_type TEXT NOT NULL DEFAULT 'student'")
        c.execute("ALTER TABLE attendance RENAME COLUMN student_id TO person_id")
    # Drop the old student_id FK constraint if it still exists (blocks satsangi inserts)
    c.execute("""SELECT constraint_name FROM information_schema.table_constraints
                 WHERE table_name='attendance' AND constraint_type='FOREIGN KEY'
                 AND constraint_name LIKE '%student%'""")
    old_fk = c.fetchone()
    if old_fk:
        c.execute(f"ALTER TABLE attendance DROP CONSTRAINT {old_fk[0]}")
    # Also drop person_id FK if it references students (same issue, different name)
    c.execute("""SELECT tc.constraint_name
                 FROM information_schema.table_constraints tc
                 JOIN information_schema.key_column_usage kcu
                   ON tc.constraint_name=kcu.constraint_name
                 JOIN information_schema.referential_constraints rc
                   ON tc.constraint_name=rc.constraint_name
                 JOIN information_schema.table_constraints tc2
                   ON rc.unique_constraint_name=tc2.constraint_name
                 WHERE tc.table_name='attendance' AND tc.constraint_type='FOREIGN KEY'
                   AND kcu.column_name='person_id'""")
    person_fk = c.fetchone()
    if person_fk:
        c.execute(f"ALTER TABLE attendance DROP CONSTRAINT {person_fk[0]}")
    c.execute('''CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)''')
    c.execute("SELECT value FROM settings WHERE key='volunteer_password'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key,value) VALUES ('volunteer_password',%s)",
                  (hash_password('volunteer123'),))
    conn.commit(); conn.close()

# ── AUTO MARK ABSENT ──────────────────────────────────────
def auto_mark_absent(event_id, event_type):
    conn = get_db(); c = conn.cursor()
    try:
        if is_sabha_type(event_type):
            c.execute("SELECT id FROM satsangis")
            person_type = 'satsangi'
        else:
            c.execute("SELECT id FROM students")
            person_type = 'student'
        all_ids = [r[0] for r in c.fetchall()]
        c.execute("SELECT person_id FROM attendance WHERE event_id=%s AND person_type=%s",
                  (event_id, person_type))
        recorded = {r[0] for r in c.fetchall()}
        absent = [(event_id, pid, person_type, 'absent') for pid in all_ids if pid not in recorded]
        if absent:
            psycopg2.extras.execute_values(c,
                "INSERT INTO attendance (event_id,person_id,person_type,status) VALUES %s ON CONFLICT DO NOTHING",
                absent)
        conn.commit()
    finally:
        conn.close()

def maybe_auto_mark_absent(event_id, event_date_str, event_type):
    """Fill absent only after the attendance window is fully closed (event date + 2 days).
    Safe to run multiple times — ON CONFLICT DO NOTHING skips existing records."""
    try:
        edate = datetime.strptime(str(event_date_str), "%Y-%m-%d").date()
        if date.today() > edate + timedelta(days=2):
            auto_mark_absent(event_id, event_type)
    except:
        pass

# ── ALLOWED ROUTES ────────────────────────────────────────
VOLUNTEER_ALLOWED = {
    'volunteer_dashboard', 'volunteer_attendance', 'volunteer_sabha_attendance',
    'volunteer_report', 'volunteer_login', 'volunteer_logout', 'volunteer_events_type',
    'api_rooms', 'api_students', 'api_mark_attendance', 'api_unmark_attendance',
    'api_search_satsangis', 'api_mark_sabha', 'api_unmark_sabha', 'api_quick_add_satsangi',
    'static'
}
PUBLIC_ROUTES = {'admin_login', 'admin_logout', 'volunteer_login', 'volunteer_logout', 'static'}

@app.before_request
def auth_guard():
    ep = request.endpoint
    if not ep or ep in PUBLIC_ROUTES: return
    is_admin = session.get('admin')
    is_volunteer = session.get('volunteer')
    if is_volunteer and not is_admin:
        if ep not in VOLUNTEER_ALLOWED:
            return redirect(url_for('volunteer_dashboard'))
        return
    if not is_admin and not is_volunteer:
        if ep not in VOLUNTEER_ALLOWED:
            return redirect(url_for('admin_login'))

# ── ADMIN AUTH ────────────────────────────────────────────
@app.route('/admin/login', methods=['GET','POST'])
def admin_login():
    error = None
    if session.get('admin'):
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        uid = request.form.get('admin_id','').strip()
        pw  = request.form.get('password','')
        if uid == ADMIN_ID and hash_password(pw) == ADMIN_PASS:
            session['admin'] = True; session.permanent = True
            return redirect(request.args.get('next') or url_for('dashboard'))
        error = 'Invalid ID or password.'
    return render_template('admin_login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('admin_login'))

# ── DASHBOARD ────────────────────────────────────────────
@app.route('/')
def dashboard():
    conn = get_db(); c = conn.cursor()
    stats = {}
    for key, table in [('buildings','buildings'),('rooms','rooms'),('students','students'),
                       ('satsangis','satsangis'),('events','events')]:
        c.execute(f'SELECT COUNT(*) FROM {table}')
        stats[key] = c.fetchone()[0]
    for etype in ['sunday', 'wednesday', 'balsabha', 'hostel']:
        c.execute("SELECT COUNT(*) FROM events WHERE event_type=%s", (etype,))
        stats[etype + '_events'] = c.fetchone()[0]
    c.execute('SELECT * FROM events ORDER BY created_at DESC LIMIT 6')
    raw_events = fetchall_dict(c)
    conn.close()
    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    recent_events = []
    for e in raw_events:
        try:
            edate = datetime.strptime(str(e['event_date']), '%Y-%m-%d').date()
            e['is_active'] = edate <= today <= edate + timedelta(days=2)
        except:
            e['is_active'] = False
        recent_events.append(e)
    return render_template('dashboard.html', stats=stats, recent_events=recent_events, today_str=today_str)

# ── SETTINGS ─────────────────────────────────────────────
@app.route('/settings/volunteer-password', methods=['POST'])
def update_volunteer_password():
    new_pw = request.form.get('new_password','')
    if len(new_pw) >= 4:
        conn = get_db(); c = conn.cursor()
        c.execute("UPDATE settings SET value=%s WHERE key='volunteer_password'", (hash_password(new_pw),))
        conn.commit(); conn.close()
    return redirect(url_for('dashboard'))

# ── VOLUNTEER AUTH ────────────────────────────────────────
@app.route('/volunteer/login', methods=['GET','POST'])
def volunteer_login():
    if session.get('volunteer') or session.get('admin'):
        return redirect(url_for('volunteer_dashboard') if session.get('volunteer') else url_for('dashboard'))
    error = None
    if request.method == 'POST':
        pw = request.form.get('password','')
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT value FROM settings WHERE key='volunteer_password'")
        stored = c.fetchone(); conn.close()
        if stored and hash_password(pw) == stored[0]:
            session['volunteer'] = True
            return redirect(url_for('volunteer_dashboard'))
        error = 'Incorrect password. Please try again.'
    return render_template('volunteer_login.html', error=error)

@app.route('/volunteer/logout')
def volunteer_logout():
    session.pop('volunteer', None)
    return redirect(url_for('volunteer_login'))

@app.route('/volunteer')
def volunteer_dashboard():
    if not session.get('volunteer') and not session.get('admin'):
        return redirect(url_for('volunteer_login'))
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT e.*, COUNT(CASE WHEN a.status='present' THEN 1 END) as attendance_count
        FROM events e LEFT JOIN attendance a ON a.event_id=e.id
        GROUP BY e.id ORDER BY e.event_date DESC""")
    all_events = fetchall_dict(c); conn.close()
    today = date.today()
    # Group active events by type
    buckets = {'sunday': [], 'wednesday': [], 'balsabha': [], 'hostel': []}
    for e in all_events:
        try:
            edate = datetime.strptime(str(e['event_date']), '%Y-%m-%d').date()
            if edate <= today <= edate + timedelta(days=2):
                e['window_open'] = True
                etype = e.get('event_type', 'hostel')
                if etype in buckets:
                    buckets[etype].append(e)
        except: pass
    return render_template('volunteer_dashboard.html', buckets=buckets)

@app.route('/volunteer/events/<etype>')
def volunteer_events_type(etype):
    if not session.get('volunteer') and not session.get('admin'):
        return redirect(url_for('volunteer_login'))
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT e.*, COUNT(CASE WHEN a.status='present' THEN 1 END) as attendance_count
        FROM events e LEFT JOIN attendance a ON a.event_id=e.id
        WHERE e.event_type=%s
        GROUP BY e.id ORDER BY e.event_date DESC""", (etype,))
    events = fetchall_dict(c)
    conn.close()
    
    today = date.today()
    active_events = []
    for e in events:
        try:
            edate = datetime.strptime(str(e['event_date']), '%Y-%m-%d').date()
            if edate <= today <= edate + timedelta(days=2):
                e['window_open'] = True
                active_events.append(e)
        except: pass

    # get event type label
    type_labels = {
        'sunday': 'Sunday Sabha',
        'wednesday': 'Wednesday Sabha',
        'balsabha': 'Bal Sabha',
        'hostel': 'Hostel Assembly'
    }
    label = type_labels.get(etype, etype.title())

    return render_template('volunteer_events_list.html', active_events=active_events, etype=etype, label=label)

@app.route('/volunteer/attendance/<int:id>')
def volunteer_attendance(id):
    if not session.get('volunteer') and not session.get('admin'):
        return redirect(url_for('volunteer_login'))
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM events WHERE id=%s', (id,))
    event = fetchone_dict(c)
    if not event: conn.close(); return redirect(url_for('volunteer_dashboard'))
    # Route to correct attendance flow based on event type
    if is_sabha_type(event.get('event_type','')):
        c.execute("SELECT person_id FROM attendance WHERE event_id=%s AND person_type='satsangi' AND status='present' ORDER BY id ASC", (id,))
        marked_ids = [r[0] for r in c.fetchall()]
        c.execute("SELECT id, name, COALESCE(mobile,'') as mobile, COALESCE(address,'') as address FROM satsangis ORDER BY name")
        all_satsangis = fetchall_dict(c)
        conn.close()
        return render_template('sabha_attendance.html', event=event, marked_ids=marked_ids,
                               all_satsangis=all_satsangis,
                               window_open=is_attendance_open(event['event_date']),
                               volunteer_mode=True)
    else:
        c.execute('SELECT * FROM buildings ORDER BY name')
        buildings = fetchall_dict(c)
        c.execute("SELECT person_id FROM attendance WHERE event_id=%s AND person_type='student' AND status='present' ORDER BY id ASC", (id,))
        marked_ids = [r[0] for r in c.fetchall()]
        c.execute('''SELECT s.id, s.name, s.roll_number, r.room_number,
               b.name as building_name, b.id as building_id, r.id as room_id
            FROM students s
            LEFT JOIN rooms r ON s.room_id=r.id
            LEFT JOIN buildings b ON r.building_id=b.id ORDER BY s.name''')
        all_students = fetchall_dict(c); conn.close()
        return render_template('attendance.html', event=event, buildings=buildings,
                               marked_ids=marked_ids, all_students=all_students,
                               window_open=is_attendance_open(event['event_date']),
                               volunteer_mode=True)

@app.route('/volunteer/report/<int:eid>')
def volunteer_report(eid):
    if not session.get('volunteer') and not session.get('admin'):
        return redirect(url_for('volunteer_login'))
    return _report(eid, volunteer_mode=True)

# ── SHARED REPORT HELPER ─────────────────────────────────
def _report(eid, volunteer_mode):
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM events WHERE id=%s', (eid,))
    event = fetchone_dict(c); conn.close()
    if not event:
        return redirect(url_for('volunteer_dashboard') if volunteer_mode else url_for('events'))
    maybe_auto_mark_absent(eid, event['event_date'], event.get('event_type', 'hostel'))
    conn = get_db(); c = conn.cursor()
    if is_sabha_type(event.get('event_type','')):
        c.execute("""SELECT s.id, s.name,
            COALESCE(s.mobile, '') as mobile,
            a.status
            FROM attendance a JOIN satsangis s ON a.person_id=s.id
            WHERE a.event_id=%s AND a.person_type='satsangi'
            ORDER BY a.status ASC, s.name""", (eid,))
        records = fetchall_dict(c)
        c.execute('SELECT COUNT(*) FROM satsangis')
    else:
        c.execute("""SELECT s.id, s.name,
            COALESCE(s.phone, '') as phone,
            COALESCE(r.room_number, '') as room_number,
            COALESCE(b.name, '') as building_name,
            a.status
            FROM attendance a JOIN students s ON a.person_id=s.id
            LEFT JOIN rooms r ON s.room_id=r.id
            LEFT JOIN buildings b ON r.building_id=b.id
            WHERE a.event_id=%s AND a.person_type='student'
            ORDER BY a.status ASC, b.name, r.room_number, s.name""", (eid,))
        records = fetchall_dict(c)
        c.execute('SELECT COUNT(*) FROM students')
    total = c.fetchone()[0]; conn.close()
    # Count present from records (accurate even before window closes)
    present_count = sum(1 for r in records if r.get('status') == 'present')
    absent_count  = total - present_count
    template = 'sabha_report.html' if is_sabha_type(event.get('event_type','')) else 'attendance_report.html'
    return render_template(template, event=event, records=records, total=total,
                           present_count=present_count, absent_count=absent_count,
                           volunteer_mode=volunteer_mode)

# ── SATSANGIS CRUD ────────────────────────────────────────
@app.route('/satsangis')
def satsangis():
    conn = get_db(); c = conn.cursor()
    search = request.args.get('q','').strip()
    if search:
        c.execute("SELECT * FROM satsangis WHERE name ILIKE %s OR COALESCE(mobile,'') ILIKE %s ORDER BY name",
                  (f'%{search}%', f'%{search}%'))
    else:
        c.execute('SELECT * FROM satsangis ORDER BY name')
    satsangis = fetchall_dict(c); conn.close()
    return render_template('satsangis.html', satsangis=satsangis, search=search)

@app.route('/satsangis/add', methods=['GET','POST'])
def add_satsangi():
    if request.method == 'POST':
        conn = get_db(); c = conn.cursor()
        c.execute('INSERT INTO satsangis (name, mobile, address) VALUES (%s,%s,%s)',
                  (request.form['name'], request.form.get('mobile',''), request.form.get('address','')))
        conn.commit(); conn.close()
        return redirect(url_for('satsangis'))
    return render_template('satsangi_form.html', satsangi=None)

@app.route('/satsangis/<int:id>/edit', methods=['GET','POST'])
def edit_satsangi(id):
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        c.execute('UPDATE satsangis SET name=%s, mobile=%s, address=%s WHERE id=%s',
                  (request.form['name'], request.form.get('mobile',''), request.form.get('address',''), id))
        conn.commit(); conn.close()
        return redirect(url_for('satsangis'))
    c.execute('SELECT * FROM satsangis WHERE id=%s', (id,))
    satsangi = fetchone_dict(c); conn.close()
    return render_template('satsangi_form.html', satsangi=satsangi)

@app.route('/satsangis/<int:id>/delete', methods=['POST'])
def delete_satsangi(id):
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM satsangis WHERE id=%s', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('satsangis'))

# ── EVENTS (combined sabha + hostel) ─────────────────────
@app.route('/events')
def events():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT e.*, COUNT(CASE WHEN a.status='present' THEN 1 END) as attendance_count
        FROM events e LEFT JOIN attendance a ON a.event_id=e.id
        GROUP BY e.id ORDER BY e.event_date DESC""")
    events_raw = fetchall_dict(c); conn.close()
    active, upcoming, past = categorize_events(events_raw)
    return render_template('events.html', active=active, upcoming=upcoming, past=past,
                           today_str=date.today().strftime('%Y-%m-%d'))

# Keep /assembly pointing to same place for backward compat
@app.route('/assembly')
def assembly():
    return redirect(url_for('events'))

@app.route('/events/add', methods=['GET','POST'])
def add_event():
    event_type = request.args.get('type', '')
    if request.method == 'POST':
        event_type = request.form.get('event_type', '')
        conn = get_db(); c = conn.cursor()
        c.execute('INSERT INTO events (title, event_date, event_type, description) VALUES (%s,%s,%s,%s)',
                  (request.form['title'], request.form['event_date'],
                   event_type, request.form.get('description','')))
        conn.commit(); conn.close()
        return redirect(url_for('events'))
    return render_template('event_form.html', event=None,
                           today=date.today().strftime('%Y-%m-%d'),
                           event_type=event_type)

@app.route('/events/<int:id>/edit', methods=['GET','POST'])
def edit_event(id):
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        c.execute('UPDATE events SET title=%s, event_date=%s, event_type=%s, description=%s WHERE id=%s',
                  (request.form['title'], request.form['event_date'],
                   request.form.get('event_type','hostel'), request.form.get('description',''), id))
        conn.commit(); conn.close()
        return redirect(url_for('events'))
    c.execute('SELECT * FROM events WHERE id=%s', (id,))
    event = fetchone_dict(c); conn.close()
    return render_template('event_form.html', event=event,
                           today=date.today().strftime('%Y-%m-%d'),
                           event_type=event.get('event_type','hostel'))

@app.route('/events/<int:id>/delete', methods=['POST'])
def delete_event(id):
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM events WHERE id=%s', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('events'))

@app.route('/events/<int:id>/attendance')
def take_attendance(id):
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM events WHERE id=%s', (id,))
    event = fetchone_dict(c)
    if is_sabha_type(event.get('event_type','')):
        c.execute("SELECT person_id FROM attendance WHERE event_id=%s AND person_type='satsangi' AND status='present' ORDER BY id ASC", (id,))
        marked_ids = [r[0] for r in c.fetchall()]
        c.execute("SELECT id, name, COALESCE(mobile,'') as mobile, COALESCE(address,'') as address FROM satsangis ORDER BY name")
        all_satsangis = fetchall_dict(c); conn.close()
        return render_template('sabha_attendance.html', event=event, marked_ids=marked_ids,
                               all_satsangis=all_satsangis,
                               window_open=is_attendance_open(event['event_date']),
                               volunteer_mode=False)
    else:
        c.execute('SELECT * FROM buildings ORDER BY name')
        buildings = fetchall_dict(c)
        c.execute("SELECT person_id FROM attendance WHERE event_id=%s AND person_type='student' AND status='present' ORDER BY id ASC", (id,))
        marked_ids = [r[0] for r in c.fetchall()]
        c.execute('''SELECT s.id, s.name, s.roll_number, r.room_number,
               b.name as building_name, b.id as building_id, r.id as room_id
            FROM students s LEFT JOIN rooms r ON s.room_id=r.id
            LEFT JOIN buildings b ON r.building_id=b.id ORDER BY s.name''')
        all_students = fetchall_dict(c); conn.close()
        return render_template('attendance.html', event=event, buildings=buildings,
                               marked_ids=marked_ids, all_students=all_students,
                               window_open=is_attendance_open(event['event_date']),
                               volunteer_mode=False)

@app.route('/events/<int:eid>/report')
def event_report(eid):
    return _report(eid, volunteer_mode=False)

# Keep old URL working
@app.route('/assembly/<int:eid>/attendance/report')
def attendance_report(eid):
    return redirect(url_for('event_report', eid=eid))

# ── BUILDINGS ─────────────────────────────────────────────
@app.route('/buildings')
def buildings():
    conn = get_db(); c = conn.cursor()
    c.execute('''SELECT b.*, COUNT(r.id) as room_count,
        (SELECT COUNT(*) FROM students s JOIN rooms r2 ON s.room_id=r2.id WHERE r2.building_id=b.id) as student_count
        FROM buildings b LEFT JOIN rooms r ON r.building_id=b.id GROUP BY b.id ORDER BY b.name''')
    buildings = fetchall_dict(c); conn.close()
    return render_template('buildings.html', buildings=buildings)

@app.route('/buildings/add', methods=['GET','POST'])
def add_building():
    if request.method == 'POST':
        conn = get_db(); c = conn.cursor()
        c.execute('INSERT INTO buildings (name, description) VALUES (%s,%s)',
                  (request.form['name'], request.form.get('description','')))
        conn.commit(); conn.close()
        return redirect(url_for('buildings'))
    return render_template('building_form.html', building=None)

@app.route('/buildings/<int:id>/edit', methods=['GET','POST'])
def edit_building(id):
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        c.execute('UPDATE buildings SET name=%s, description=%s WHERE id=%s',
                  (request.form['name'], request.form.get('description',''), id))
        conn.commit(); conn.close()
        return redirect(url_for('buildings'))
    c.execute('SELECT * FROM buildings WHERE id=%s', (id,))
    building = fetchone_dict(c); conn.close()
    return render_template('building_form.html', building=building)

@app.route('/buildings/<int:id>/delete', methods=['POST'])
def delete_building(id):
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM buildings WHERE id=%s', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('buildings'))

# ── ROOMS ─────────────────────────────────────────────────
@app.route('/buildings/<int:bid>/rooms')
def rooms(bid):
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM buildings WHERE id=%s', (bid,))
    building = fetchone_dict(c)
    c.execute('''SELECT r.*, COUNT(s.id) as occupancy
        FROM rooms r LEFT JOIN students s ON s.room_id=r.id
        WHERE r.building_id=%s GROUP BY r.id ORDER BY r.floor, r.room_number''', (bid,))
    rooms = fetchall_dict(c); conn.close()
    return render_template('rooms.html', rooms=rooms, building=building)

@app.route('/buildings/<int:bid>/rooms/add', methods=['GET','POST'])
def add_room(bid):
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        c.execute('INSERT INTO rooms (building_id, room_number, capacity, floor) VALUES (%s,%s,%s,%s)',
                  (bid, request.form['room_number'], request.form.get('capacity',4), request.form.get('floor',1)))
        conn.commit(); conn.close()
        return redirect(url_for('rooms', bid=bid))
    c.execute('SELECT * FROM buildings WHERE id=%s', (bid,))
    building = fetchone_dict(c); conn.close()
    return render_template('room_form.html', room=None, building=building)

@app.route('/rooms/<int:id>/edit', methods=['GET','POST'])
def edit_room(id):
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        c.execute('UPDATE rooms SET room_number=%s, capacity=%s, floor=%s WHERE id=%s',
                  (request.form['room_number'], request.form.get('capacity',4), request.form.get('floor',1), id))
        conn.commit()
        c.execute('SELECT * FROM rooms WHERE id=%s', (id,))
        room = fetchone_dict(c); conn.close()
        return redirect(url_for('rooms', bid=room['building_id']))
    c.execute('SELECT r.*, b.name as building_name FROM rooms r JOIN buildings b ON r.building_id=b.id WHERE r.id=%s', (id,))
    room = fetchone_dict(c)
    c.execute('SELECT * FROM buildings WHERE id=%s', (room['building_id'],))
    building = fetchone_dict(c); conn.close()
    return render_template('room_form.html', room=room, building=building)

@app.route('/rooms/<int:id>/delete', methods=['POST'])
def delete_room(id):
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT building_id FROM rooms WHERE id=%s', (id,))
    bid = c.fetchone()[0]
    c.execute('UPDATE students SET room_id=NULL WHERE room_id=%s', (id,))
    c.execute('DELETE FROM rooms WHERE id=%s', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('rooms', bid=bid))

# ── STUDENTS ──────────────────────────────────────────────
@app.route('/students')
def students():
    conn = get_db(); c = conn.cursor()
    c.execute('''SELECT s.*, r.room_number, b.name as building_name
        FROM students s LEFT JOIN rooms r ON s.room_id=r.id
        LEFT JOIN buildings b ON r.building_id=b.id ORDER BY s.name''')
    students = fetchall_dict(c); conn.close()
    return render_template('students.html', students=students)

@app.route('/students/add', methods=['GET','POST'])
def add_student():
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        room_id = request.form.get('room_id') or None
        error = None
        if room_id:
            c.execute('SELECT capacity, COUNT(s.id) as occ FROM rooms r LEFT JOIN students s ON s.room_id=r.id WHERE r.id=%s GROUP BY r.id', (room_id,))
            row = c.fetchone()
            if row and row[1] >= row[0]:
                error = f'Room is full (capacity {row[0]}). Please choose another room or increase capacity.'
        if error:
            c.execute('SELECT * FROM buildings ORDER BY name')
            buildings = fetchall_dict(c)
            c.execute('SELECT r.*, b.name as bname FROM rooms r JOIN buildings b ON r.building_id=b.id ORDER BY b.name, r.room_number')
            rooms = fetchall_dict(c); conn.close()
            return render_template('student_form.html', student=None, buildings=buildings, rooms=rooms, error=error)
        c.execute('INSERT INTO students (name, roll_number, phone, room_id, joining_date) VALUES (%s,%s,%s,%s,%s)',
                  (request.form['name'], request.form.get('roll_number',''),
                   request.form.get('phone',''), room_id,
                   request.form.get('joining_date', date.today().strftime('%Y-%m-%d'))))
        conn.commit(); conn.close()
        return redirect(url_for('students'))
    c.execute('SELECT * FROM buildings ORDER BY name')
    buildings = fetchall_dict(c)
    c.execute('SELECT r.*, b.name as bname FROM rooms r JOIN buildings b ON r.building_id=b.id ORDER BY b.name, r.room_number')
    rooms = fetchall_dict(c); conn.close()
    return render_template('student_form.html', student=None, buildings=buildings, rooms=rooms)

@app.route('/students/<int:id>/edit', methods=['GET','POST'])
def edit_student(id):
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        room_id = request.form.get('room_id') or None
        error = None
        if room_id:
            c.execute('SELECT capacity, COUNT(s.id) as occ FROM rooms r LEFT JOIN students s ON s.room_id=r.id AND s.id!=%s WHERE r.id=%s GROUP BY r.id', (id, room_id))
            row = c.fetchone()
            if row and row[1] >= row[0]:
                error = f'Room is full (capacity {row[0]}). Please choose another room or increase capacity.'
        if error:
            c.execute('SELECT * FROM students WHERE id=%s', (id,))
            student = fetchone_dict(c)
            c.execute('SELECT * FROM buildings ORDER BY name')
            buildings = fetchall_dict(c)
            c.execute('SELECT r.*, b.name as bname FROM rooms r JOIN buildings b ON r.building_id=b.id ORDER BY b.name, r.room_number')
            rooms = fetchall_dict(c); conn.close()
            return render_template('student_form.html', student=student, buildings=buildings, rooms=rooms, error=error)
        c.execute('UPDATE students SET name=%s, roll_number=%s, phone=%s, room_id=%s, joining_date=%s WHERE id=%s',
                  (request.form['name'], request.form.get('roll_number',''),
                   request.form.get('phone',''), room_id,
                   request.form.get('joining_date',''), id))
        conn.commit(); conn.close()
        return redirect(url_for('students'))
    c.execute('SELECT * FROM students WHERE id=%s', (id,))
    student = fetchone_dict(c)
    c.execute('SELECT * FROM buildings ORDER BY name')
    buildings = fetchall_dict(c)
    c.execute('SELECT r.*, b.name as bname FROM rooms r JOIN buildings b ON r.building_id=b.id ORDER BY b.name, r.room_number')
    rooms = fetchall_dict(c); conn.close()
    return render_template('student_form.html', student=student, buildings=buildings, rooms=rooms)

@app.route('/students/<int:id>/delete', methods=['POST'])
def delete_student(id):
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM students WHERE id=%s', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('students'))

# ── API ───────────────────────────────────────────────────
@app.route('/api/buildings/<int:bid>/rooms')
def api_rooms(bid):
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id, room_number, floor FROM rooms WHERE building_id=%s ORDER BY floor, room_number', (bid,))
    rooms = fetchall_dict(c); conn.close()
    return jsonify(rooms)

@app.route('/api/rooms/<int:rid>/students')
def api_students(rid):
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id, name, roll_number FROM students WHERE room_id=%s ORDER BY name', (rid,))
    students = fetchall_dict(c); conn.close()
    return jsonify(students)

@app.route('/api/attendance/mark', methods=['POST'])
def api_mark_attendance():
    data = request.get_json()
    conn = get_db(); c = conn.cursor()
    try:
        c.execute('SELECT event_date FROM events WHERE id=%s', (data['event_id'],))
        event = c.fetchone()
        if not event or not is_attendance_open(event[0]):
            conn.close()
            return jsonify({'success': False, 'error': 'Attendance window is closed.'})
        c.execute('''INSERT INTO attendance (event_id, person_id, person_type, status)
            VALUES (%s,%s,'student','present') ON CONFLICT DO NOTHING''',
                  (data['event_id'], data['student_id']))
        conn.commit()
        c.execute("SELECT COUNT(*) FROM attendance WHERE event_id=%s AND person_type='student' AND status='present'", (data['event_id'],))
        count = c.fetchone()[0]; conn.close()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/attendance/unmark', methods=['POST'])
def api_unmark_attendance():
    data = request.get_json()
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM attendance WHERE event_id=%s AND person_id=%s AND person_type='student'",
              (data['event_id'], data['student_id']))
    conn.commit()
    c.execute("SELECT COUNT(*) FROM attendance WHERE event_id=%s AND person_type='student' AND status='present'", (data['event_id'],))
    count = c.fetchone()[0]; conn.close()
    return jsonify({'success': True, 'count': count})

@app.route('/api/satsangis/search')
def api_search_satsangis():
    q = request.args.get('q','').strip()
    if len(q) < 1:
        return jsonify([])
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id, name, COALESCE(mobile,'') as mobile FROM satsangis WHERE name ILIKE %s OR COALESCE(mobile,'') ILIKE %s ORDER BY name LIMIT 20",
              (f'%{q}%', f'%{q}%'))
    results = fetchall_dict(c); conn.close()
    return jsonify(results)

@app.route('/api/sabha/mark', methods=['POST'])
def api_mark_sabha():
    data = request.get_json()
    conn = get_db(); c = conn.cursor()
    try:
        c.execute('SELECT event_date FROM events WHERE id=%s', (data['event_id'],))
        event = c.fetchone()
        if not event or not is_attendance_open(event[0]):
            conn.close()
            return jsonify({'success': False, 'error': 'Attendance window is closed.'})
        c.execute('''INSERT INTO attendance (event_id, person_id, person_type, status)
            VALUES (%s,%s,'satsangi','present') ON CONFLICT DO NOTHING''',
                  (data['event_id'], data['satsangi_id']))
        conn.commit()
        c.execute("SELECT COUNT(*) FROM attendance WHERE event_id=%s AND person_type='satsangi' AND status='present'", (data['event_id'],))
        count = c.fetchone()[0]; conn.close()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/sabha/unmark', methods=['POST'])
def api_unmark_sabha():
    data = request.get_json()
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM attendance WHERE event_id=%s AND person_id=%s AND person_type='satsangi'",
              (data['event_id'], data['satsangi_id']))
    conn.commit()
    c.execute("SELECT COUNT(*) FROM attendance WHERE event_id=%s AND person_type='satsangi' AND status='present'", (data['event_id'],))
    count = c.fetchone()[0]; conn.close()
    return jsonify({'success': True, 'count': count})


# ── ANALYTICS ─────────────────────────────────────────────
@app.route('/analytics')
def analytics():
    conn = get_db(); c = conn.cursor()

    # ── Trend per event type ──────────────────────────────
    trends = {}
    event_counts = {}
    for etype in ['sunday', 'wednesday', 'balsabha', 'hostel']:
        ptype = 'student' if etype == 'hostel' else 'satsangi'
        c.execute("""
            SELECT e.event_date, e.title,
                COUNT(CASE WHEN a.status='present' THEN 1 END) as present_count,
                COUNT(CASE WHEN a.status='absent'  THEN 1 END) as absent_count
            FROM events e
            LEFT JOIN attendance a ON a.event_id=e.id AND a.person_type=%s
            WHERE e.event_type=%s
            GROUP BY e.id ORDER BY e.event_date ASC
        """, (ptype, etype))
        trends[etype] = fetchall_dict(c)
        c.execute("SELECT COUNT(*) FROM events WHERE event_type=%s", (etype,))
        event_counts[etype] = c.fetchone()[0]

    # ── Per-student hostel stats ──────────────────────────
    c.execute("""
        SELECT s.id, s.name, COALESCE(s.roll_number,'') as roll_number,
            COALESCE(r.room_number,'') as room_number,
            COALESCE(b.name,'') as building_name,
            COUNT(CASE WHEN a.status='present' THEN 1 END) as present_count,
            COUNT(CASE WHEN a.status='absent'  THEN 1 END) as absent_count
        FROM students s
        LEFT JOIN rooms r ON s.room_id=r.id
        LEFT JOIN buildings b ON r.building_id=b.id
        LEFT JOIN attendance a ON a.person_id=s.id AND a.person_type='student'
        GROUP BY s.id, s.name, s.roll_number, r.room_number, b.name
        ORDER BY present_count DESC, s.name
    """)
    student_stats = fetchall_dict(c)

    # ── Satsangi regularity analysis ─────────────────────
    # For each satsangi: how many times present in each sabha type
    c.execute("""
        SELECT s.id, s.name, COALESCE(s.mobile,'') as mobile,
            COUNT(CASE WHEN e.event_type='sunday'    AND a.status='present' THEN 1 END) as sun_present,
            COUNT(CASE WHEN e.event_type='wednesday' AND a.status='present' THEN 1 END) as wed_present,
            COUNT(CASE WHEN e.event_type='balsabha'  AND a.status='present' THEN 1 END) as bal_present,
            COUNT(CASE WHEN e.event_type='sunday'    THEN 1 END) as sun_total,
            COUNT(CASE WHEN e.event_type='wednesday' THEN 1 END) as wed_total,
            COUNT(CASE WHEN e.event_type='balsabha'  THEN 1 END) as bal_total
        FROM satsangis s
        LEFT JOIN attendance a ON a.person_id=s.id AND a.person_type='satsangi'
        LEFT JOIN events e ON a.event_id=e.id
        GROUP BY s.id, s.name, s.mobile
        ORDER BY (
            COUNT(CASE WHEN a.status='present' THEN 1 END)
        ) DESC, s.name
    """)
    satsangi_stats = fetchall_dict(c)

    # Classify each satsangi by regularity pattern
    THRESHOLD = 0.6  # 60%+ attendance = "regular" for that type
    for s in satsangi_stats:
        types = []
        for key, total_key, label in [
            ('sun_present','sun_total','Sunday'),
            ('wed_present','wed_total','Wednesday'),
            ('bal_present','bal_total','Bal Sabha'),
        ]:
            total = s[total_key]
            present = s[key]
            if total > 0 and present / total >= THRESHOLD:
                types.append(label)
        if len(types) == 0:
            s['regularity'] = 'Irregular'
            s['reg_class'] = 'irregular'
        elif len(types) == 3:
            s['regularity'] = 'All Sabha'
            s['reg_class'] = 'all'
        elif len(types) == 1:
            s['regularity'] = f'{types[0]} Regular'
            s['reg_class'] = types[0].lower().replace(' ','')
        else:
            s['regularity'] = ' + '.join(types)
            s['reg_class'] = 'multi'
        s['total_present'] = s['sun_present'] + s['wed_present'] + s['bal_present']
        s['total_events']  = s['sun_total']   + s['wed_total']   + s['bal_total']

    conn.close()
    return render_template('analytics.html',
                           trends=trends,
                           event_counts=event_counts,
                           student_stats=student_stats,
                           satsangi_stats=satsangi_stats)

# ── QUICK ADD SATSANGI (from attendance page) ──────────────
@app.route('/api/satsangis/quick-add', methods=['POST'])
def api_quick_add_satsangi():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    mobile = (data.get('mobile') or '').strip()
    address = (data.get('address') or '').strip()
    if not name or not mobile:
        return jsonify({'success': False, 'error': 'Name and mobile are required'})
    conn = get_db(); c = conn.cursor()
    try:
        c.execute('INSERT INTO satsangis (name, mobile, address) VALUES (%s,%s,%s) RETURNING id',
                  (name, mobile, address or None))
        new_id = c.fetchone()[0]
        conn.commit(); conn.close()
        return jsonify({'success': True, 'id': new_id, 'name': name, 'mobile': mobile, 'address': address})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

# ── IMPORT ────────────────────────────────────────────────
@app.route('/import', methods=['GET','POST'])
def import_data():
    result = None
    if request.method == 'POST':
        import_type = request.form.get('import_type', 'hostel')
        f = request.files.get('datafile')
        if f and f.filename:
            try:
                import openpyxl, io, tempfile, os as _os
                data = f.read()
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(data); tmp_path = tmp.name
                if import_type == 'satsangi':
                    result = _import_satsangis(tmp_path)
                else:
                    result = _import_hostel(tmp_path)
                _os.unlink(tmp_path)
            except Exception as e:
                result = {'status': 'error', 'message': str(e), 'inserted': 0, 'skipped': 0}
        else:
            result = {'status': 'error', 'message': 'No file selected.', 'inserted': 0, 'skipped': 0}
    return render_template('import_data.html', result=result)

def _import_satsangis(filepath):
    """Import satsangis from single-column Excel (column A = name)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    inserted = skipped = 0
    conn = get_db(); c = conn.cursor()
    for row in ws.iter_rows(min_row=1):
        val = row[0].value
        if not val: continue
        name = str(val).strip()
        if not name or name.lower() in ('name','full name','satsangi','sr','sr.'): continue
        # Skip if already exists
        c.execute("SELECT id FROM satsangis WHERE name ILIKE %s", (name,))
        if c.fetchone():
            skipped += 1; continue
        c.execute("INSERT INTO satsangis (name) VALUES (%s)", (name,))
        inserted += 1
    conn.commit(); conn.close()
    return {'status': 'success', 'inserted': inserted, 'skipped': skipped,
            'message': f'Satsangi import complete.'}

def _import_hostel(filepath):
    """Import hostel data from Excel with merged-cell building names."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    # Find merged rows = building names
    merged_rows = {}
    for mc in ws.merged_cells.ranges:
        cell = ws.cell(mc.min_row, mc.min_col)
        if cell.value:
            merged_rows[mc.min_row] = str(cell.value).strip()
    # Parse rows
    current_building = None
    skip_next = False
    students = []
    for row_num in range(1, ws.max_row + 1):
        if row_num in merged_rows:
            current_building = merged_rows[row_num]
            skip_next = True; continue
        if skip_next:
            skip_next = False; continue
        row = ws[row_num]
        vals = [c.value for c in row]
        if len(vals) < 2: continue
        room, name = vals[0], vals[1]
        if not name or not isinstance(name, str) or not name.strip(): continue
        name = name.strip()
        if name.lower() in ('full name','name','total',''): continue
        if not room: continue
        room_str = str(room).strip()
        if room_str.endswith('.0'): room_str = room_str[:-2]
        if not current_building: continue
        students.append({'building': current_building, 'room': room_str, 'name': name})
    if not students:
        return {'status': 'error', 'message': 'No student data found in file.', 'inserted': 0, 'skipped': 0}
    # Insert into DB
    conn = get_db(); c = conn.cursor()
    building_ids = {}; room_ids = {}
    inserted = skipped = 0
    for s in students:
        bname = s['building']
        rname = s['room']
        # Get or create building
        if bname not in building_ids:
            c.execute("SELECT id FROM buildings WHERE name ILIKE %s", (bname,))
            row = c.fetchone()
            if row:
                building_ids[bname] = row[0]
            else:
                c.execute("INSERT INTO buildings (name) VALUES (%s) RETURNING id", (bname,))
                building_ids[bname] = c.fetchone()[0]
        bid = building_ids[bname]
        # Get or create room
        room_key = f"{bid}:{rname}"
        if room_key not in room_ids:
            c.execute("SELECT id FROM rooms WHERE building_id=%s AND room_number=%s", (bid, rname))
            row = c.fetchone()
            if row:
                room_ids[room_key] = row[0]
            else:
                c.execute("INSERT INTO rooms (building_id, room_number, capacity, floor) VALUES (%s,%s,%s,%s) RETURNING id",
                          (bid, rname, 10, 1))
                room_ids[room_key] = c.fetchone()[0]
        rid = room_ids[room_key]
        # Insert student if not duplicate name in same building
        c.execute("""SELECT s.id FROM students s JOIN rooms r ON s.room_id=r.id
                     WHERE r.building_id=%s AND s.name ILIKE %s""", (bid, s['name']))
        if c.fetchone():
            skipped += 1; continue
        c.execute("INSERT INTO students (name, room_id) VALUES (%s,%s)", (s['name'], rid))
        inserted += 1
    conn.commit(); conn.close()
    return {'status': 'success', 'inserted': inserted, 'skipped': skipped,
            'message': f'Hostel import complete. {len(set(s["building"] for s in students))} buildings processed.'}

# ── STARTUP ───────────────────────────────────────────────
init_db()

if __name__ == '__main__':
    app.run(debug=False)